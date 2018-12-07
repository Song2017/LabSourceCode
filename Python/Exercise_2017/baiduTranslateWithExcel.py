# /usr/bin/env python
# coding=utf8

import hashlib
import random
import openpyxl
from openpyxl import Workbook
import requests

# set baidu develop parameter
apiurl = 'http://api.fanyi.baidu.com/api/trans/vip/translate'
appid = '20180816000194959'
secretKey = 'lWzwUiWmhRORknf68FCT'

# 翻译内容 源语言 翻译后的语言
def translateBaidu(content, fromLang='en', toLang='zh'):
    salt = str(random.randint(32768, 65536))
    sign = appid + content + salt + secretKey
    sign = hashlib.md5(sign.encode("utf-8")).hexdigest()

    try:
        paramas = {
            'appid': appid,
            'q': content,
            'from': fromLang,
            'to': toLang,
            'salt': salt,
            'sign': sign
        }
        response = requests.get(apiurl, paramas)
        jsonResponse = response.json()  # 获得返回的结果，结果为json格式
        dst = str(jsonResponse["trans_result"]
                  [0]["dst"])  # 取得翻译后的文本结果
        return dst
    except Exception as e:
        print(e)


def excelTrans(srcFilename=r'c:\_Work\source.xlsx', desFilename=r'c:\_Work\result.xlsx'
               srcSheet='Sheet2',  srcColumn=1, srcRowBegin=1, srcRowEnd=28, 
               desColumn=1, desSheet='result2'):
    wb = openpyxl.load_workbook(srcFilename)
    ws = wb[srcSheet]
    wb2 = Workbook()
    ws2 = wb2.create_sheet(title=desSheet)

    for i in range(srcRowBegin, srcRowEnd, 1):
        result = ws.cell(row=i, column=srcColumn).value
        if not (result is None):
            ws2.cell(row=i-srcRowBegin+1,
                     column=desColumn).value = translateBaidu(result)

    wb2.save(desFilename)


if __name__ == '__main__':
    print('translate begin...')
    excelTrans()
    print('ending...')